"""
风险分析和管理总表 Excel 生成服务

根据参考文件 KF-SAP1-02-0003 风险分析和管理总表.xlsm 的样式，
生成 17 列 (A~Q) 两行表头的风险矩阵 Excel (.xlsx)：

- product_risk_analysis_matrix        -> 风险分析和管理总表
  (中文, 严重度×发生概率 S*O 双栏评估：控制前/剩余)
- cybersecurity_risk_analysis_matrix  -> 网络安全风险分析和管理总表
  (中英双语, 严重度/发生概率/可探测度 S/O/D 双栏评估：Before/After Control)

参考文件为 .xlsm（含 VBA 宏），openpyxl 无法保留宏，故输出 .xlsx。
"""

import json
import re
from typing import List, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.services.minimax import _call_minimax_api_raw


# ── 常量：样式 ───────────────────────────────────────────────────────────────

_THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
_HEADER_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
_TITLE_FONT = Font(name="宋体", size=14, bold=True)
_HEADER_FONT = Font(name="宋体", size=10, bold=True)
_DATA_FONT = Font(name="宋体", size=9)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


# ── 格式定义 ─────────────────────────────────────────────────────────────────

# 产品风险分析总表（中文 S×O）
_PRODUCT_RISK = {
    "sheet_name": "风险分析和管理总表",
    "title": "风险分析和管理总表",
    "row2": [
        "条目", "危害分类", "可预见的事件序列", "危险情况", "伤害(失效的后果)",
        "严重度\nSeverity", "发生概率\nprobability", "风险系数R=S*O", "风险等级\nRisk level",
        "控制措施", "剩余风险评估", "", "", "",
        "控制措施验证结果", "是否还存在剩余风险", "剩余风险可接受的理由",
    ],
    "row3": [
        "", "", "", "", "",
        "", "", "", "",
        "", "严重度\nSeverity", "发生概率\nprobability", "风险系数", "风险等级",
        "", "", "",
    ],
    # 表头横向跨列合并 (如 A2:A3 纵向、K2:N2 横向)
    "merges": [
        "A2:A3", "B2:B3", "C2:C3", "D2:D3", "E2:E3", "F2:F3", "G2:G3",
        "H2:H3", "I2:I3", "J2:J3", "K2:N2", "O2:O3", "P2:P3", "Q2:Q3",
    ],
    "widths": {
        "A": 8.2, "B": 12.5, "C": 29.4, "D": 21.6, "E": 23.2, "F": 9.1,
        "G": 8.9, "H": 6.2, "I": 7.6, "J": 27.8, "K": 9.6, "L": 9.9,
        "M": 5.8, "N": 6.8, "O": 33.1, "P": 6.8, "Q": 13.9,
    },
    "fields": [
        "item", "hazard_classification", "foreseeable_event", "hazardous_situation", "harm",
        "severity_before", "probability_before", "risk_score", "risk_level",
        "control_measures",
        "severity_after", "probability_after", "risk_score_after", "risk_level_after",
        "verification", "residual_risk", "acceptability",
    ],
}

# 网络安全风险分析总表（中英双语 S/O/D）
_CYBER_RISK = {
    "sheet_name": "网络安全风险分析和管理总表",
    "title": "网络安全风险分析和管理总表",
    "row2": [
        "Item", "Hazard Classification", "Potential Failure Causation",
        "Potential Failure Mode", "Hazard Description",
        "Before Control", "", "", "",
        "Current Measures",
        "After Control", "", "", "",
        "Requirement traceability",
        "Design traceability(only to software Control Measures)",
        "Verification results of Control Measure",
    ],
    "row3": [
        "", "", "", "", "",
        "Severity", "probability", "Detectability", "Risk level",
        "",
        "Severity", "probability", "Detectability", "Risk level",
        "", "", "",
    ],
    "merges": [
        "A2:A3", "B2:B3", "C2:C3", "D2:D3", "E2:E3", "F2:I2", "J2:J3",
        "K2:N2", "O2:O3", "P2:P3", "Q2:Q3",
    ],
    "widths": {
        "A": 6.0, "B": 10.0, "C": 20.8, "D": 18.2, "E": 41.1, "F": 11.0,
        "G": 11.0, "H": 13.0, "I": 9.0, "J": 57.2, "K": 11.0, "L": 11.0,
        "M": 13.0, "N": 9.0, "O": 16.2, "P": 16.1, "Q": 32.0,
    },
    "fields": [
        "item", "hazard_classification", "failure_causation", "failure_mode", "hazard_description",
        "severity_before", "probability_before", "detectability_before", "risk_level_before",
        "current_measures",
        "severity_after", "probability_after", "detectability_after", "risk_level_after",
        "requirement_traceability", "design_traceability", "verification_results",
    ],
}

_RISK_FORMATS: Dict[str, dict] = {
    "product_risk_analysis_matrix": _PRODUCT_RISK,
    "cybersecurity_risk_analysis_matrix": _CYBER_RISK,
}


def is_risk_matrix_doc(doc_type: str) -> bool:
    """判断文档类型是否为风险分析总表类（可导出 Excel）。"""
    return doc_type in _RISK_FORMATS


# ── LLM 结构化风险数据生成 ───────────────────────────────────────────────────

_RISK_SYSTEM_PROMPT = """你是专业的医疗器械风险管理文档撰写助手。请根据产品信息生成风险分析和管理总表的结构化风险条目数据。

严格要求：
1. 只输出一个 JSON 数组，不要输出任何解释、前言、markdown 围栏或后缀
2. 数组每个元素是一个风险条目对象，字段键严格遵循给定的字段清单（缺一不可）
3. 风险条目覆盖主要危害类别，数量 5~8 条，编号连续
4. 内容专业、具体，符合 ISO 14971 / GB/T 42062（网络安全类则 YY/T 1843）风险分析要求
5. 所有字段值均为字符串；风险系数(risk_score)填整数（用字符串表示，如 "9"）
"""

# 字段说明（用于拼进 user prompt）
_PRODUCT_FIELD_DESC = {
    "item": "条目编号，如 3.01、3.02",
    "hazard_classification": "危害分类：信息危害/操作(运行)危害/生物化学危害/电气危害/机械危害/软件危害 等",
    "foreseeable_event": "可预见的事件序列",
    "hazardous_situation": "危险情况",
    "harm": "伤害(失效的后果)",
    "severity_before": "控制前严重度，如「严重的 3」「中等的 2」「轻度的 1」",
    "probability_before": "控制前发生概率，如「偶然发生 3」「极少发生 2」「难以置信 1」",
    "risk_score": "风险系数 R=S*O 的整数结果",
    "risk_level": "控制前风险等级：低 / ALARP / 不可接受",
    "control_measures": "控制措施",
    "severity_after": "剩余严重度",
    "probability_after": "剩余发生概率",
    "risk_score_after": "剩余风险系数（整数）",
    "risk_level_after": "剩余风险等级：ACC / ALARP",
    "verification": "控制措施验证结果",
    "residual_risk": "是否还存在剩余风险：是 / 否",
    "acceptability": "剩余风险可接受的理由，无则填 /",
}

_CYBER_FIELD_DESC = {
    "item": "条目编号，如 4.2.1",
    "hazard_classification": "危害分类，如「Operational hazards 操作危害」",
    "failure_causation": "Potential Failure Causation 潜在失效原因（中英）",
    "failure_mode": "Potential Failure Mode 潜在失效模式（中英）",
    "hazard_description": "Hazard Description 危害描述（中英）",
    "severity_before": "Before Control 严重度，如「Serious 3」「Negligible 1」",
    "probability_before": "Before Control 概率，如「Improbable 2」「Incredible 1」",
    "detectability_before": "Before Control 可探测度，如「Undetectable 5」「Detectable 3」",
    "risk_level_before": "Before Control 风险等级：ACC / ALARP",
    "current_measures": "Current Measures 现有控制措施（中英）",
    "severity_after": "After Control 严重度",
    "probability_after": "After Control 概率",
    "detectability_after": "After Control 可探测度",
    "risk_level_after": "After Control 风险等级：ACC / ALARP",
    "requirement_traceability": "Requirement traceability 需求追溯：文档号+条款号",
    "design_traceability": "Design traceability 设计追溯：文档号+章节号",
    "verification_results": "Verification results 验证结果：测试报告编号",
}


def _extract_json_array(text: str) -> List[dict]:
    """从 LLM 返回文本中提取 JSON 数组（容错 markdown 围栏 / 前后缀）。"""
    if not text:
        return []
    m = re.search(r"```(?:json)?\s*\n(.*?)```", text, re.DOTALL)
    if m:
        text = m.group(1)
    start = text.find("[")
    end = text.rfind("]")
    if start == -1 or end == -1 or end <= start:
        return []
    try:
        data = json.loads(text[start:end + 1])
    except json.JSONDecodeError:
        return []
    if isinstance(data, list):
        return [x for x in data if isinstance(x, dict)]
    if isinstance(data, dict):
        for key in ("rows", "risks", "items", "data"):
            if isinstance(data.get(key), list):
                return [x for x in data[key] if isinstance(x, dict)]
        return [data]
    return []


def generate_risk_rows(
    doc_type: str,
    product_name: str = "",
    classification: str = "",
    intended_use: str = "",
) -> List[dict]:
    """调用本地 Ollama 生成结构化风险条目数据，返回 list[dict]。

    每个 dict 以 17 字段键与 build_risk_excel 的 fields 对齐。
    失败或模型未返回有效 JSON 时返回空列表。
    """
    fmt = _RISK_FORMATS.get(doc_type)
    if not fmt:
        return []
    field_desc = (
        _CYBER_FIELD_DESC
        if doc_type == "cybersecurity_risk_analysis_matrix"
        else _PRODUCT_FIELD_DESC
    )

    lines = []
    for key, desc in field_desc.items():
        lines.append(f'  "{key}": {desc}')
    field_spec = "\n".join(lines)

    user_prompt = (
        f"产品名称：{product_name or '贴敷式胰岛素泵'}\n"
        f"产品分类：{classification or 'Ⅱ类有源医疗器械'}\n"
        f"预期用途：{intended_use or '用于糖尿病患者的胰岛素持续皮下输注'}\n\n"
        f"请按以下字段清单生成 JSON 数组（每个字段必填）：\n"
        f"{field_spec}\n"
    )

    result = _call_minimax_api_raw(
        _RISK_SYSTEM_PROMPT,
        user_prompt,
        temperature=0.3,
        max_tokens=16384,
        timeout=(30, 180),
    )
    return _extract_json_array(result or "")


# ── Excel 构建 ───────────────────────────────────────────────────────────────

def build_risk_excel(doc_type: str, rows: List[dict]) -> bytes:
    """根据格式定义构建风险矩阵 .xlsx，返回文件字节。

    Args:
        doc_type: product_risk_analysis_matrix 或 cybersecurity_risk_analysis_matrix
        rows: 风险条目列表，每个元素为含 17 字段键的 dict
    """
    fmt = _RISK_FORMATS.get(doc_type)
    if not fmt:
        raise ValueError(f"不支持的风险矩阵文档类型: {doc_type}")

    wb = Workbook()
    ws = wb.active
    ws.title = fmt["sheet_name"]

    # 1) 标题行：A1:Q1 合并
    ws.merge_cells("A1:Q1")
    ws["A1"] = fmt["title"]
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = _CENTER
    ws.row_dimensions[1].height = 30

    # 2) 两行表头 (row 2 / row 3)
    for idx, val in enumerate(fmt["row2"], start=1):
        c = ws.cell(row=2, column=idx, value=val if val else None)
        c.font = _HEADER_FONT
        c.alignment = _CENTER
        c.fill = _HEADER_FILL
        c.border = _THIN_BORDER
    for idx, val in enumerate(fmt["row3"], start=1):
        c = ws.cell(row=3, column=idx, value=val if val else None)
        c.font = _HEADER_FONT
        c.alignment = _CENTER
        c.fill = _HEADER_FILL
        c.border = _THIN_BORDER
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 24

    # 3) 表头合并单元格
    for merge in fmt["merges"]:
        ws.merge_cells(merge)

    # 4) 数据行（从第 4 行开始）
    fields = fmt["fields"]
    start_row = 4
    for r_idx, row in enumerate(rows):
        excel_row = start_row + r_idx
        for col_idx, key in enumerate(fields, start=1):
            val = row.get(key, "")
            if val is None:
                val = ""
            # 风险系数字段：尽力转 int 以便数值对齐
            if key in ("risk_score", "risk_score_after") and isinstance(val, str):
                try:
                    val = int(val)
                except (ValueError, TypeError):
                    pass
            c = ws.cell(row=excel_row, column=col_idx, value=val)
            c.font = _DATA_FONT
            c.border = _THIN_BORDER
            if key in ("item", "risk_score", "risk_score_after", "residual_risk",
                       "severity_before", "probability_before", "risk_level",
                       "severity_after", "probability_after", "risk_level_after",
                       "detectability_before", "detectability_after",
                       "risk_level_before"):
                c.alignment = _CENTER
            else:
                c.alignment = _LEFT

    # 5) 列宽
    for letter, width in fmt["widths"].items():
        ws.column_dimensions[letter].width = width

    # 6) 写出字节
    import io
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
